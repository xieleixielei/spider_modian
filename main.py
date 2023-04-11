# -*- codeing = utf-8 -*-
import random
import re  # 正则表达式，进行文字匹配`
import socket
import time
import urllib.error  # 制定URL，获取网页数据
import urllib.request
import requests
import xlwt  # 进行excel操作
from bs4 import BeautifulSoup  # 网页解析，获取数据
# from get_detail_3_30 import get_data,get_author
import spiders_re as my_re
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

findLink = re.compile(r'<a href="(.*?)"')
findImgSrc = re.compile(r'<img.*src="(.*?)"', re.S)
findTitle = re.compile(r'<h3 class="pro_title">(.*)</h3>')
findmoney = re.compile(r'<span backer_money="(\d+)">(.*)</span>')
findrate = re.compile(r'rate="(\d+)">(.*)</span>')
findgoal = re.compile(r'<span class="goal-money">目标金额 (.*)</span>')
findsponsor = re.compile(r'<span backer_count="(\d+)">(.*)</span>')

# detail page
find_starttime = re.compile(r'start_time="(.*?)"')
find_starttime2 = re.compile(r'<h3>(.*?)</h3>')

find_endtime = re.compile(r'end_time="(.*?)"')
find_remaintime = re.compile(r'<h3(.*?)\b(start_time|end_time)\b(.*?)>(.*?)</h3>')

savepath = "摩点众筹all.xls"  # 当前目录新建XLS，存储进去
idea_key = 0
going_key = 0
preheat_key = 0
success_key = 0
fail_key=0
global sheet
global book
global itemreal_class
datalist = []  # 用来存储爬取的网页信息
excelindex = 0
not_fail_id=[]
baseurl_list = [
                # "https://zhongchou.modian.com/all/top_time/success/",
                # "https://zhongchou.modian.com/all/top_time/going/",
                # "https://zhongchou.modian.com/all/top_time/preheat/",
                "https://zhongchou.modian.com/all/top_time/idea/"
                ]
def main():
    global book_all, update_sheet, author_sheet
    book_all = Workbook()
    update_sheet = book_all.create_sheet(title='update')
    author_sheet = book_all.create_sheet(title='author')
    # 添加表头
    header = ['项目id', '次数', '时间', '内容', '图片', '系统更新', '评论数', '点赞', '抽奖', '开奖']
    update_sheet.append(header)
    row = ['项目id', '作者名称', '星标', '星标tag', '支持项目数', '头像图片']
    author_sheet.append(row)

    global book
    book = xlwt.Workbook(encoding="utf-8", style_compression=0)  # 创建workbook对象
    global success_key, going_key, preheat_key, idea_key, fail_key, sheet
    # sheet = book.add_sheet('摩点众筹爬取-preheat', cell_overwrite_ok=True)  # 创建工作表
    # sheet = book.add_sheet('摩点众筹爬取-idea', cell_overwrite_ok=True)  # 创建工作表
    # sheet = book.add_sheet('摩点众筹爬取-going', cell_overwrite_ok=True)  # 创建工作表
    # sheet = book.add_sheet('摩点众筹爬取-success', cell_overwrite_ok=True)  # 创建工作表
    sheet = book.add_sheet('all',cell_overwrite_ok=True)
    global datalist,excelindex
    datalist = []  # 用来存储爬取的网页信息
    excelindex = 0
    getData("https://zhongchou.modian.com/all/top_time/all/")
    # for url in baseurl_list:
    #     getData(url)
    saveData()
    book.save(savepath)  # 保存


def get_authorapi(uid):

    home_url = "https://apim.modian.com/apis/comm/user/user_info"
    params = {"json_type": 1, "to_user_id": uid, "user_id": uid}
    # "timestamp": int(time.time())
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 10; SM-G975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Mobile Safari/537.36',
        'Origin': 'https://m.modian.com',
        'Referer': 'https://m.modian.com/',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Timestamp': str(int(time.time()))  # 添加时间戳参数
    }
    response = requests.get(home_url, params=params,headers=headers)
    print(response.json())
    pass


def get_authorinfo(html,author_id):
    fans_num=0
    notice_number=0
    love_number=0
    detail_result={}
    other_result={}
    data=[]
    soup3 = BeautifulSoup(html, "html.parser")
    # print(soup3)
    for items in soup3.find_all('div', {'class': 'banner'}):  # 查找符合要求的字符串
        # print(items)
        userId_flg = items.find('div', {'class': 'userId'})
        if userId_flg:
            userId = userId_flg.text.strip()
        else:
            userId = "none"
        # print("userId",userId)
        for item in items.find_all('div', {'class': 'cont'}):  # 查找符合要求的字符串
            fans_item = item.find('span', {'class': 'go_span fans'})
            # print("fan_items",fans_item)
            fans_item=str(fans_item)
            try:
                fans_num=re.findall(my_re.find_fannum,fans_item)[0]
            except IndexError:
                fans_num=0
            print(fans_num)

            notice_item = item.select_one('span.go_span:not(.fans)')
            # print('notice_item',notice_item)
            if notice_item:
                text = notice_item.text.strip()  # 获取标签中的文本内容，去掉两端的空白字符
                notice_number = text.split()[0]  # 使用空格分割文本内容，并取第一个元素作为数字
            else:
                notice_number=0
            print(notice_number)

            add_item = item.find('span', {'id': 'ALL'})
            if add_item:
                text2 = add_item.text.strip()  # 获取标签中的文本内容，去掉两端的空白字符
                love_number = text2.split()[0]  # 使用空格分割文本内容，并取第一个元素作为数字
            else:
                love_number=0
            print(love_number)
    detail_result = {}
    for itemss in soup3.find_all('div', {'class': 'detail'}):  # 查找符合要求的字符串
        item_s = itemss.find_all('div', class_='item')
        for itt in item_s:
            detail_label = itt.find('label').text
            detail_text = itt.find('p').text
            detail_result[detail_label] = detail_text
        # print(detail_result)
    other_result = {}
    for itemm in soup3.find_all('div', {'class': 'other_info'}):  # 查找符合要求的字符串
        otheritems = itemm.find_all('div', class_='item')
        for item_other in otheritems:
            value = int(item_other.find('p').text)
            key = item_other.find_all('p')[1].text
            other_result[key] = value
        # print(other_result)
    '''
        第四层网页
        用户主页
        # 格式:https://m.modian.com/user/homePage/id
        api :
    '''
    userhome_baseurl="https://m.modian.com/user/homePage/"
    userhome_url=userhome_baseurl+str(author_id)
    print('url',userhome_url)
    # get_authorapi(author_id)

    # homehtml=askURL2(userhome_url)
    # soup4 = BeautifulSoup(homehtml, "html.parser")
    # print(soup4)
    # for user_items in soup4.find_all('div', {'class': 'header'}):
    #     # print('dad',user_items)
    #     pass
    '''
    fans_num notice_number love_number detail_result other_result userhome_url
    '''
    data.append(fans_num)
    data.append(notice_number)
    data.append(love_number)
    data.append(str(detail_result))
    data.append(str(other_result))
    data.append(userhome_url)
    # print("data is:",data)
    return data
    # list_tmp=[]
    # return list_tmp

def get_upperitems(soup2):
    global true_authorid, sponsor_info
    global itemreal_class
    data = []
    global idea_key,preheat_key,success_key,going_key
    # time info
    if preheat_key == 1:
        for item in soup2.find_all('div', {'class': 'col2 start-time'}):
            item = str(item)
            # print(item)
            try:
                starttime = re.findall(find_starttime2, item)[0]
                # endtime = "预热中"
            except IndexError:
                starttime="none"
                # endtime="none"
            # print(starttime)
            try:
                endtime = re.findall(find_endtime, item)[0]
            except IndexError:
                endtime="预热中"
            if starttime !="none":
                import datetime
                now = datetime.datetime.now()
                print("starttime",starttime)
                try:
                    starttime_std = datetime.datetime.strptime(starttime, "%Y-%m-%d %H:%M")
                    if starttime_std < now:
                        print("love")
                        (success_key, going_key, preheat_key, idea_key) = (0, 1, 0, 0)
                        itemreal_class = "众筹中"
                        endtime="众筹中——异常值"
                except ValueError:
                    pass
            data.append(starttime)
            data.append(endtime)
    elif idea_key == 1:
        starttime = "创意中"
        endtime = "创意中"
        print(starttime)
        print(endtime)
        data.append(starttime)
        data.append(endtime)
    else:
        for item in soup2.find_all('div', {'class': 'col2 remain-time'}):  # 查找符合要求的字符串
            item = str(item)
            print(item)
            try:
                starttime = re.findall(find_starttime, item)[0]
                endtime = re.findall(find_endtime, item)[0]
            except IndexError:
                starttime="none"
                endtime="none"
            # print(starttime, endtime)
            data.append(starttime)
            data.append(endtime)
    data.append(itemreal_class)
    # author info
    for items in soup2.find_all('div', {'class': 'sponsor-info clearfix'}):  # 查找符合要求的字符串
        # sponsor-link
        # print('new-items:\n',items)
        # 使用find_all方法查找所有<a>标签
        a_tags = items.find_all('a')
        for a_tag in a_tags:
            sponsor_href = a_tag.get('href')
            print("sponsor-link", sponsor_href)
            # find author id
            sponsor_href_str=str(sponsor_href)
            try:
                true_authorid=re.findall(my_re.find_trueauthorid,sponsor_href_str)[0]
            except IndexError:
                true_authorid="none"
            print('userid',true_authorid)
            detail_html = askURL(sponsor_href)  # 保存获取到的网页源码
            sponsor_info=get_authorinfo(detail_html,true_authorid)
            # data.append(true_authorid)
            # data.extend(sponsor_info)
            data.append(sponsor_href)
        # sponsor-image
        items = str(items)
        try:
            author_image = re.findall(my_re.find_authorimage, items)[0]
        except IndexError:
            author_image="none"
        # print("author_image", author_image)
        data.append(author_image)
        try:
            catagory = re.findall(my_re.find_catagory, items)[0]
        except IndexError:
            catagory="none"
        # print("catagory", catagory)
        data.append(catagory)
        try:
            author_name = re.findall(my_re.find_authorname, items)[0]
        except IndexError:
                author_name="none"
        # print("name:", author_name)
        data.append(author_name)
        try:
            author_uid = re.findall(my_re.find_authoruid, items)[0]
        except IndexError:
            author_uid = 0
        # print("uid:", author_uid)
        data.append(author_uid)
    # project base info
    if preheat_key == 1:
        money = 0
        percent = 0
        data.append(money)
        data.append(percent)
        for items in soup2.find_all('div', {'class': 'center'}):
            items_str = str(items)
            # print(items_str)
            for item in items.find_all('div', {'class': 'col1 project-goal'}):
                item = str(item)
                # print(item)
                try:
                    goal_money = re.findall(my_re.find_preheatgoal, item)[0]
                except IndexError:
                    goal_money=0
                # print("goal", goal_money)
                data.append(goal_money)
            try:
                sponsor_content = re.findall(my_re.findsubscribe, items_str)[0]
                sponsor = sponsor_content[1]
            except IndexError:
                sponsor="none"
            # print("sponsor", sponsor)
            data.append(sponsor)
    elif idea_key == 1:
        money = 0
        percent = 0
        goal_money = 'none'
        sponsor = 'none'
        data.append(money)
        data.append(percent)
        data.append(goal_money)
        data.append(sponsor)
    else:
        for items in soup2.find_all('div', {'class': 'center'}):  # 查找符合要求的字符串
            # print('new-items:\n',items)
            items = str(items)
            # money
            try:
                content = re.findall(findmoney, items)[0]
                money = content[1]
            except IndexError:
                money =0
            # print('目前已筹集￥',money)
            data.append(money)
            # rate
            try:
                rate_content = re.findall(findrate, items)[0]
                percent = rate_content[1]
            except IndexError:
                percent=0
            # print('当前进度为',percent)
            data.append(percent)
            # goal
            try:
                goal_money = re.findall(findgoal, items)[0]
            except IndexError:
                goal_money=0
            # print("goal", goal_money)
            data.append(goal_money)
            # sponsor
            try:
                sponsor_content = re.findall(findsponsor, items)[0]
                sponsor_num = sponsor_content[1]
            except IndexError:
                sponsor_num=0
            print("当前支持人数为%s人" % sponsor_num)
            data.append(sponsor_num)
    data.append(true_authorid)
    data.extend(sponsor_info)
    return data


def get_main_left(soup2):
    data = []
    # print("hello")
    # for items in soup2.find_all('ul', {'class': 'tack-lists'}):  # 查找符合要求的字符串
    #     print("hello")
    #     print(items)
    # print("hello")
    for items in soup2.find_all('div', {'class': 'main-left'}):  # 查找符合要求的字符串
        # print(items)
        for item in items.find_all('div', {'class': 'project-content'}):  # 查找符合要求的字符串
            img_list = []
            video_list = []
            # 提取所有 img 标签中的图片
            for img in item.find_all('img'):
                img_src = img.get('src')
                if img_src:
                    img_list.append(img_src)
            # 提取所有 video 标签中的视频
            for video in item.find_all('video'):
                # for source in video.find_all('source'):
                    video_src = video.get('src')
                    if video_src:
                        video_list.append(video_src)
            # print(f'共找到 {len(img_list)} 个图片：{img_list}')
            # print(f'共找到 {len(video_list)} 个视频：{video_list}')
            data.append(len(img_list))
            data.append(str(img_list))
            data.append(len(video_list))
            data.append(str(video_list))

    return data


def get_main_right(soup2):
    # backlist
    global backtitle, backtime, backtext
    data = []
    test_data=[]
    for items in soup2.find_all('div', {'class': 'main-right'}):  # 查找符合要求的字符串
      # print(items)
      for item in items.find_all('div', {'class': 'payback-lists margin36'}):  # 查找符合要求的字符串
          for elements in item.find_all('div', class_=lambda x: x and 'back-list' in x):
              sub_datalist=[]
              # print("---")
              # print(elements)
              element = elements.find_all('div', {'class': 'head'})
              element=str(element)
              # print(element)
              try:
                  backmoney = re.findall(my_re.find_backheadmoney, element)[0]
                  backsponsor_content = re.findall(my_re.find_backheadsponsor, element)[0]
                  backsponsor = backsponsor_content[1]
                  backsponsor=backsponsor.replace('\n', '')
                  backsponsor=backsponsor.strip()
              except IndexError:
                  backmoney = 0
                  backsponsor=0
              # print(backsponsor)
              # print(backmoney)

              sign_logo_elements=elements.find_all('div', {'class': 'zc-subhead'})
              sign_logo_elements=str(sign_logo_elements)
              # print("-----++++++\n",sign_logo_elements)
              try:
                  sign_logo = re.findall(my_re.find_backheadsignlogo,sign_logo_elements)[0]
              except IndexError:
                  sign_logo = 0# bug
              # print(sign_logo)
              for content_items in elements.find_all('div', {'class': 'back-content'}):
                  content_item = content_items.find('div', {'class': 'back-sub-title'})
                  content_item=str(content_item)
                  # print(content_item)
                  try:
                      backtitle=re.findall(my_re.find_backlisttilte,content_item)[0]
                  except  IndexError:
                      backtitle = "none"
                  # print(backtitle)
                  backtext = content_items.find('div', {'class': 'back-detail'}).text.strip()
                  # print(backtext)
                  backtime_flg=content_items.find('div', {'class': 'back-time'})
                  if backtime_flg:
                      backtime = backtime_flg.text.strip()
                  else:
                      backtime="none"
                  # print(backtime)
              sub_datalist.append(backtitle)
              sub_datalist.append(sign_logo)
              sub_datalist.append(backmoney)
              sub_datalist.append(backsponsor)
              sub_datalist.append(backtime)
              sub_datalist.append(backtext)
              # print("sub_data:",sub_datalist)
              test_data.append(str(sub_datalist))
    '''
        data list add 
        [[backtitle,signlogo,backmoney,backsponsor,backtime,backtext]]
    '''
    data.append(str(test_data))
    num=len(test_data)
    data.append(num)
    # print(test_data)
    return data



global user_num_pass
def get_main_middle(soup2):
    data = []
    userlist_num=0
    collect_number=0
    update_number=0
    comment_number=0
    for items in soup2.find_all('div', {'class': 'nav-wrap-inner'}):  # 查找符合要求的字符串
        # print("dbaicid",items)
        for item in items.find_all('ul', {'class': 'nav-left'}):  # 查找符合要求的字符串
            # print("dbaicid",item)
            update_time_item= item.find_all('li',{ 'class':'pro-gengxin'})
            update_time_item=str(update_time_item)
            try:
                update_time_item_content=re.findall(my_re.find_update_time_item,update_time_item)[0]
                update_number=update_time_item_content[1]
            except IndexError:
                update_number=0
            if update_number =="":
                update_number=0
            print("update___:",update_number)

            comment_item= item.find_all('li',{ 'class':'nav-comment'})
            comment_item=str(comment_item)
            try:
                comment_item_content=re.findall(my_re.find_comment_item,comment_item)[0]
                comment_number=comment_item_content[1]
            except IndexError:
                comment_number=0
            if comment_number =="":
                comment_number=0
            print("comment___:",comment_number)

            userlist_item = item.find_all('li', class_='dialog_user_list')
            userlist_item = str(userlist_item)
            # print(userlist_item)
            if idea_key == 1:
                try:
                    userlist_content = re.findall(my_re.find_idea_userlist_item, userlist_item)[0]
                    userlist_num = userlist_content[1]
                except IndexError:
                    userlist_num=0
                if userlist_num == "":
                    userlist_num=0
                print("userlist_num___:", userlist_num)
            else:
                try:
                    userlist_content = re.findall(my_re.find_userlist_item,userlist_item)[0]
                    userlist_num = userlist_content[1]
                except IndexError:
                    userlist_num=0
                if userlist_num == "":
                    userlist_num=0
                print("userlist_num___:",userlist_num)

        if idea_key == 1:
            collect_number=userlist_num
        else:
            for item in items.find_all('ul', {'class': 'nav-right'}):  # 查找符合要求的字符串
                collect_item= item.find_all('li',{ 'class':'atten'})
                collect_item=str(collect_item)
                try:
                    collect_number=re.findall(my_re.find_collect_item,collect_item)[0]
                except:
                    collect_number=0
                if collect_number == "":
                    collect_number=0
        print("collect_number___:",collect_number)
    global user_num_pass
    user_num_pass=userlist_num
    data.append(update_number)
    data.append(comment_number)
    data.append(userlist_num)
    data.append(collect_number)
    return data


global book_all,update_sheet,author_sheet

# def pass_main_get():
#     global book_all, update_sheet, author_sheet
#     global get_detail_id,user_num_pass,idea_key
#     data1 = get_data(get_detail_id)
#     ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
#     # text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
#     page_num_q = int(int(user_num_pass) / 20) + 1
#     data2 = get_author(get_detail_id, idea_key, page_num_q)
#     if data1 is not None:
#         for data_row in data1:
#             # update_sheet.append(data_row)
#             data_row = [ILLEGAL_CHARACTERS_RE.sub(r'', (str(cell))) for cell in data_row]
#             try:
#                 update_sheet.append(data_row)
#             except IllegalCharacterError as e:
#                 print(f"Invalid character found: {e}")
#
#     if data2 is not None:
#         for author_row in data2:
#             # author_sheet.append(author_row)
#             clean_row = [ILLEGAL_CHARACTERS_RE.sub(r'', (str(cell))) for cell in author_row]
#             try:
#                 author_sheet.append(clean_row)
#             except IllegalCharacterError as e:
#                 print(f"Invalid character found: {e}")
#     # print(data2)
#     book_all.save('items_all_3_31_19_31.xlsx')



def get_downitems(soup2):
    data = []
    mainright_items = get_main_right(soup2)
    data.extend(mainright_items)

    mainmiddle_items=get_main_middle(soup2)
    data.extend(mainmiddle_items)

    # pass_main_get()

    mainleft_items = get_main_left(soup2)
    data.extend(mainleft_items)
    return data

def get_class(class_result):
    global idea_key,preheat_key,going_key,success_key
    item_class = ""
    if class_result == "看好":
        item_class="创意"
        idea_key=1
    elif class_result == "看好项目":
        item_class="预热"
        preheat_key=1
    elif class_result == "立即购买支持":
        item_class="众筹中"
        going_key=1
    elif class_result == "众筹成功":
        item_class="众筹成功"
        success_key=1
    elif class_result == "项目终止":
        item_class="项目终止"
        success_key=1
    elif class_result == "众筹结束":
        item_class="众筹失败"
        going_key = 1
    elif class_result == "众筹取消":
        item_class = "众筹取消"
        going_key = 1
    else :
        item_class="未知情况"
    return  item_class



def get_detail(html):
    data = []
    global preheat_key, idea_key
    soup2 = BeautifulSoup(html, "html.parser")

    result_button = soup2.find_all('div', {'class': 'buttons clearfloat'})  # 查找符合要求的字符串
    result_button=result_button[0]
    result_class=result_button.select_one('.buttons.clearfloat > a').text.strip()
    print(result_class)
    global itemreal_class
    itemreal_class=get_class(result_class)
    # print(soup2)

    upperitems = get_upperitems(soup2)
    data.extend(upperitems)
    downitems = get_downitems(soup2)
    data.extend(downitems)

    return data

global get_detail_id

# 爬取网页
def getData(baseurl):
    global datalist, book,excelindex,get_detail_id
    int_project = 0
    page = 0
    # page_num = get_page()
    for i in range(1, 833):  # 调用获取页面信息的函数，497次
    # for i in range(1, page_num+1):
    # for i in range(1, 2):  # 调试调用获取页面信息的函数，1次
        url = baseurl + str(i)
        html = askURL(url)  # 保存获取到的网页源码
        # 2.逐一解析数据
        soup = BeautifulSoup(html, "html.parser")
        # print(soup)
        item_index = 0
        for items in soup.find_all('div', {'class': 'pro_field'}):  # 查找符合要求的字符串
            # print(items)
            for item in items.find_all('li'):
                # 调试
                # if item_index == 1:
                #     break
                item_index = item_index + 1
                excelindex = excelindex + 1
                global idea_key,going_key,preheat_key,success_key
                idea_key=0
                going_key=0
                preheat_key=0
                success_key=0
                print(
                    "----------------------------------------============%d=============----------------------------------------" % item_index)
                # print(item)
                data = []  # 保存一个项目的所有信息
                item = str(item)
                data.append(excelindex)
                # ---------parser start--------
                # 0. link
                link = re.findall(findLink, item)[0]  # 通过正则表达式查找
                print(link)
                find_linkid = re.compile(r'https://zhongchou.modian.com/item/(\d+).html')
                try:
                    linkid = re.findall(find_linkid, str(link))[0]
                except IndexError:
                    linkid = ""
                link="https://zhongchou.modian.com/item/"+str(linkid)+".html"
                print(linkid)
                data.append(link)
                data.append(linkid)
                get_detail_id=linkid
                # 1.project-title
                titles = re.findall(findTitle, item)[0]
                if "可汗游戏大会" in titles:
                    continue
                print(titles)
                data.append(titles)

                # 2.project-img
                imgSrc = re.findall(findImgSrc, item)[0]
                print(imgSrc)
                data.append(imgSrc)
                # -----enter detail page----
                detail_html = askURL(link)  # 保存获取到的网页源码
                with open('example.html', 'w', encoding='utf-8') as f:
                    f.write(detail_html)
                detail_data = get_detail(detail_html)
                data.extend(detail_data)
                datalist.append(data)
                # 每条数据一写，每1页一存
                saveData()
                datalist=[]
                if i%2==0:
                    book.save(savepath)
        int_project = int_project + item_index
        page = page + 1
        print("page number is", page)
        print("this page has %d project\n" % item_index)
    print('总数为', int_project)
    # print("datalist is ", datalist)


# 得到指定一个URL的网页内容
def askURL(url):
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 Edg/111.0.1661.44"
    }

    # 设置超时时间的范围
    timeout_range = (5, 15)
    request = urllib.request.Request(url, headers=head)
    html = ""
    for i in range(3):
        try:
            # 生成随机超时时间
            timeout = random.randint(*timeout_range)
            response = urllib.request.urlopen(request, timeout=timeout)
            html = response.read().decode("utf-8")
            break
        # except urllib.error.URLError as e:
        except (urllib.error.URLError, ConnectionResetError, socket.timeout) as e:
            if i == 9:  # 如果已经是最后一次重试，直接退出循环
                saveData()
                print('重试多次仍然失败！')
                break
            else:
                print(f'第{i + 1}次尝试失败，原因：{e}')
                time.sleep(0.3)  # 等待 1 秒后重试
    return html


# 得到指定一个URL的网页内容
def askURL2(url):
    head = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 10; SM-G975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Mobile Safari/537.36'
    }

    # 设置超时时间的范围
    timeout_range = (5, 15)
    request = urllib.request.Request(url, headers=head)
    html = ""
    for i in range(3):
        try:
            # 生成随机超时时间
            timeout = random.randint(*timeout_range)
            response = urllib.request.urlopen(request, timeout=timeout)
            html = response.read().decode("utf-8")
            break
        # except urllib.error.URLError as e:
        except (urllib.error.URLError, ConnectionResetError, socket.timeout) as e:
            if i == 9:  # 如果已经是最后一次重试，直接退出循环
                saveData()
                print('重试多次仍然失败！')
                break
            else:
                print(f'第{i + 1}次尝试失败，原因：{e}')
                time.sleep(0.3)  # 等待 1 秒后重试
    return html

# 保存数据到表格
def saveData():
    global sheet, book, savepath,excelindex
    print("save.......")
    # fans_num notice_number love_number detail_result other_result userhome_url
    col = (
        "序号",
        "项目link","项目6位id","项目名称", "项目图","开始时间", "结束时间","项目结果",  "用户主页","用户头像", "分类",
        "用户名", "项目id", "已筹金额","百分比", "目标金额", "支持者","uid","粉丝数","关注数","赞数",
        "发起人信息列表","发起人发起项目信息","发起人详细主页","回报列表信息-标题-限量-金额-标签-数量-内容", "回报列表项目数",
        "项目更新数","评论数","项目支持者列表人数","收藏数",
        "项目详情-图片数量", "项目详情-图片", "项目详情-视频数量","项目详情-视频"
    )
    for i in range(0, 34):
        sheet.write(0, i, col[i])  # 列名

    save_length = len(datalist)
    # for i in range(0, save_length):
    # for i in range(excel, save_length):
        # print(datalist)
    try:
        data = datalist[0]
    except IndexError:
        data=[]
        # print("data length", len(data))
        # 覆盖原数据
    for j in range(0, 34):
        if len(data) > j:
            data_string=str(data[j])
            # 判断字符串的长度是否超过 32767 个字符
            if len(data_string) > 32767:
                # 将超长字符串分割成多个小字符串
                max_length = 32767
                sub_strings = [data_string[i:i + max_length] for i in range(0, len(data_string), max_length)]
                # 将每个小字符串写入到不同的单元格中
                times=0
                for i, sub_string in enumerate(sub_strings):
                    sheet.write(excelindex, j, sub_string)
                    times=times+1
                j=j+times-1
            else:
                sheet.write(excelindex, j, data[j])
        else:
            sheet.write(excelindex, j, "None")
            # sheet.write(i + 1, j, data[j])  # 数据
    # book.save(savepath)  # 保存


if __name__ == "__main__":  # 当程序执行时
    main()
    print("爬取完毕！")
